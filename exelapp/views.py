from django.shortcuts import render
from django.shortcuts import render
from django.http import HttpResponse
from .models import Order
import os
import pandas as pd
import pandas as df
from openpyxl import Workbook
from exelapp.models import Order


# Create your views here.
def index(request):
  return render(request,"app/upload.html")

def upload_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        # Save the uploaded Excel file on the server
        save_excel_file(excel_file)
        # Parse the uploaded Excel file and insert its data into the database
        parse_excel_and_insert_data(excel_file)
        return HttpResponse('Upload successful')
    return render(request, 'upload.html')
from django.http import HttpResponse

def download_excel(request):
    # Logic to retrieve data from the database and generate the Excel file
    # ...
    # Return the Excel file as a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=output.xlsx'
    workbook.save(response)
    return response

def save_excel_file(excel_file):
    import os

def save_excel_file(excel_file):
    # Specify the directory where you want to save the file
    upload_dir = 'C:/Users/Shubham/OneDrive/Desktop/exel project/exelapp/exelfile/output.xlsx'

    # Create the upload directory if it doesn't exist
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir)

    # Generate a unique file name or use the original file name
    file_name = excel_file.name  # Use the original file name
    # If you want to generate a unique file name, you can use a combination of datetime and a unique identifier

    # Construct the file path
    file_path = os.path.join(upload_dir, file_name)

    # Save the file to the specified location
    with open(file_path, 'wb') as destination:
        for chunk in excel_file.chunks():
            destination.write(chunk)

    return file_path


def parse_excel_and_insert_data(excel_file):


 
 df = pd.read_excel('C:/Users/Shubham/OneDrive/Desktop/exel project/exelapp/exelfile/output.xlsx')

# Assuming you have already set up a database connection and have a table called 'orders'

# Iterate over each row in the DataFrame and insert data into the database table
import pandas as pd
from openpyxl import Workbook

# Assuming you have already set up a database connection and have a table called 'orders'

# Fetch the data from the database table and store it in a pandas DataFrame
from exelapp.models import Order

queryset = Order.objects.all()
df = pd.DataFrame(list(queryset.values()))

# Create a new Excel workbook
workbook = Workbook()

# Select the active sheet
sheet = workbook.active

# Write the column headers to the first row of the sheet
for col_num, column_header in enumerate(df.columns, 1):
    sheet.cell(row=1, column=col_num, value=column_header)

# Write the data from the DataFrame to the sheet, starting from the second row
for row in df.itertuples(index=False):
    row_num = sheet.max_row + 1
    for col_num, cell_value in enumerate(row, 1):
        sheet.cell(row=row_num, column=col_num, value=cell_value)

# Save the workbook to a file
output_file_path = 'C:/Users/Shubham/OneDrive/Desktop/exel project/exelapp/exelfile/output.xlsx'

workbook.save(output_file_path)
